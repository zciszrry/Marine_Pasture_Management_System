import requests
from django.shortcuts import render,HttpResponse,redirect
from app01.models import Staff,Account
from .utils.error import *
import hashlib
from .utils import getHomeData
from app01.models import Device,HydrologicInfo,Fish,NetCage,FishGroup,Sensor
import json
import psutil
from django.http import JsonResponse
from django.db.models import Sum,Count
from django.core.paginator import Paginator
from app01.utils.code import check_code
from io import BytesIO
from openpyxl import load_workbook,Workbook
from datetime import datetime,timedelta
from scipy.stats import gaussian_kde
import numpy as np
from collections import defaultdict
import pandas as pd
import logging
# Create your views here.

# 默认参数 request

def login(request):
    if request.method == 'GET':
        return render(request,'login.html')
    else:
        uname=request.POST.get('username')
        pwd=request.POST.get('password')
        code=request.POST.get('code')
        # 如果超时为空
        image_code=request.session.get('image_code',"")
        if image_code.upper() != code.upper():
            # print(image_code,code)
            return errorResponse(request, '验证码错误')

        md5 = hashlib.md5()
        md5.update(pwd.encode())
        pwd = md5.hexdigest()

        try:
            user=Account.objects.get(username=uname,password=pwd)
            print(user)
            staff=Staff.objects.get(id=user.id)
            print(staff)
            if staff.state == 'A':
                return errorResponse(request, '该员工状态异常')

            request.session['username']=user.username
            # 登陆成功后，session保存7天
            request.session.set_expiry(60*60*24*7)
            return redirect('/app01/home')
        except:
            return errorResponse(request, '用户名或密码错误')


def image_code(request):
    """生成图片验证码"""
    img,code_str=check_code()
    # print(code_str)
    # 写入session，以便后续获取验证码进行校验
    request.session['image_code']=code_str
    # 设置session60秒超时，验证码过期
    request.session.set_expiry(60)

    stream = BytesIO()
    img.save(stream, 'png')

    return HttpResponse(stream.getvalue())

def registry(request):
    if request.method=='GET':
        return render(request,'registry.html')
    else:
        uname=request.POST.get('username')
        staff_id=request.POST.get('staff_id')
        pwd=request.POST.get('password')
        # print(uname,staff_id,pwd)
        try:
            Staff.objects.get(id=staff_id)
        except:
            return errorResponse(request,'非本公司职员无法注册')
        try:
            Account.objects.get(username=uname)
        except:
            md5=hashlib.md5()
            md5.update(pwd.encode())
            pwd=md5.hexdigest()
            Account.objects.create(id=int(staff_id),username=uname,password=pwd)
            return redirect('/app01/login')

        return errorResponse(request, '该用户名已经被注册')

def control(request):
    # 获取所有信息
    data_list=Staff.objects.all()
    # print(data_list)
    return render(request,"control.html",{"data_list":data_list})

def logout(request):
    request.session.clear()
    return redirect('login')

def home(request):
    uname=request.session.get('username')
    user=Account.objects.get(username=uname)
    staffinfo=Staff.objects.get(id=user.id)
    y,m,d=getHomeData.getNowTime()
    return render(request,'index.html',{
        'staffInfo':staffinfo,
        'userInfo':user,
        'dateInfo':{'year':y,'month':m,'day':d},
    })


# Create your views here.

def data(request):
    # 首先按category分组，然后按device_id排序
    devices = Device.objects.all().order_by('category')
    print(devices)
    # 初始化一个字典来保存每个类别的当前排序ID
    category = None
    id=1
    # 为devices附加类别内的排序id信息
    data_sum=0
    for device in devices:
        if category == None:
            category = device.category
            device.category_index_id = id
            id+=1
        elif category == device.category:
            device.category_index_id = id
            id+=1
        else:
            category = device.category
            device.category_index_id = 1
            id=2
        data_sum+=device.memory

    paginator = Paginator(devices, 5)  # 每页显示设备
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "data.html", {'page_obj': page_obj,'data_sum': data_sum})


def get_data_status(request):
    type_memory_stats = Device.objects.values('type').annotate(total_memory=Sum('memory')).order_by('type')
    type_memory_name_value = [(item['type'], item['total_memory']) for item in type_memory_stats]

    # 提取键列表和值列表
    type_names = [item[0] for item in type_memory_name_value]  # 键列表，存储类型名称
    memory_values = [item[1] for item in type_memory_name_value]  # 值列表，存储内存总量

    # 打印结果，以验证
    print("类型名称列表:", type_names)
    print("内存总量列表:", memory_values)
    data = {
        'type_names': type_names,
        'memory_values': memory_values
    }

    # 返回 JsonResponse，包含键列表和值列表
    return JsonResponse(data)

def get_hardware_status(request):
    # 获取CPU使用率
    cpu_usage = psutil.cpu_percent(interval=1)

    # 获取内存信息
    memory = psutil.virtual_memory()
    memory_usage = memory.percent

    # 获取磁盘分区信息
    disk_usage = psutil.disk_usage('/')
    disk_usage_percent = disk_usage.percent

    # 构建响应数据
    hardware_status = {
        'cpu_usage': cpu_usage,
        'memory_usage': memory_usage,
        'disk_usage': disk_usage_percent,
    }

    # 将数据转换为JSON格式并返回
    return JsonResponse(hardware_status)

def info(request):
    return render(request, "main_info.html" )

def center(request):
    return render(request, "intelligence_center.html" )

def water(request):
    # 水质信息+环境评分（最新）
    latest_hydrologic_info = HydrologicInfo.objects.latest('monitoring_time')
    latest_environment_score = latest_hydrologic_info.calculate_environment_score()
    rounded_score = round(latest_environment_score)

    # 绘制总数-时间折线
    fish_groups = FishGroup.objects.all().order_by('time')
    fish_summary = defaultdict(int)
    for fg in fish_groups:
        date_str = fg.time.strftime('%Y-%m-%d')
        fish_summary[date_str] += int(fg.number)

    # 转换为前端需要的格式
    fish_data = [{'time': date, 'number': number} for date, number in sorted(fish_summary.items())]

    diff=0
    # 假设 fish_data 已经按照上面的代码片段生成
    if len(fish_data) >= 2:
        # 访问列表的最后两个元素
        last_item = fish_data[-1]  # 最后一个元素
        second_last_item = fish_data[-2]  # 倒数第二个元素
        # 计算差值
        diff = last_item['number'] - second_last_item['number']
    add_num=0
    die_num=0
    if diff > 0:
       add_num=diff
    else:
        die_num=-diff


    # 鱼群各种类数量
    fish_groups = FishGroup.objects.values('type').annotate(total_number=Sum('number'))
    fish_data_type = [{'type': fg['type'], 'total_number': fg['total_number']} for fg in fish_groups]

    # 传感器
    NetCage_data =NetCage.objects.first()
    Sensor_data =Sensor.objects.first()

    # 鱼群重量信息
    fish_weights = list(Fish.objects.values_list('weight', flat=True))

    # 使用核密度估计 (Kernel Density Estimation) 计算概率密度函数
    kde = gaussian_kde(fish_weights)
    weight_min = min(fish_weights)
    weight_max = max(fish_weights)
    x_values = np.linspace(weight_min, weight_max, 1000)
    density_values = kde(x_values)

    # 归一化使得面积为1
    density_values /= np.trapz(density_values, x_values)

    # 计算鱼群总数
    total_fish_groups = FishGroup.objects.count()

    # 计算鱼种类总数
    fish_species = Fish.objects.values('type').distinct()
    total_fish_species = fish_species.count()

    # 计算总设备数
    total_devices = Device.objects.count()

    # 查询不同设备种类的数量
    category_counts = Device.objects.values('category').annotate(count=Count('category'))
    # 获取设备种类的总数
    total_categories = category_counts.count()

    # 查询所有鱼群数据，按时间顺序排列
    fish_groups = FishGroup.objects.order_by('type', '-time')

    # 初始化一个字典来存储每种鱼群类型对应的最新数量和时间
    fish_type_dict = {}

    # 遍历每条鱼群数据，更新每种鱼群类型对应的最新数量和时间
    for fish_group in fish_groups:
        if fish_group.type in fish_type_dict:
            # 如果鱼类型在字典中，比较日期并更新数量和时间
            if fish_group.time > fish_type_dict[fish_group.type]['time']:
                fish_type_dict[fish_group.type] = {
                    'number': int(fish_group.number),
                    'time': fish_group.time
                }
        else:
            # 如果鱼类型不在字典中，添加类型和数量到字典中
            fish_type_dict[fish_group.type] = {
                'number': int(fish_group.number),
                'time': fish_group.time
            }

    # 求和所有鱼群类型对应的最新数量
    total_latest_fish = sum(item['number'] for item in fish_type_dict.values())

    context = {
        "latest_hydrologic_info": latest_hydrologic_info,
        "latest_environment_score": rounded_score,
        'fish_data': fish_data,
        'fish_data_type': fish_data_type,
        'netcage_data': NetCage_data,
        'sensor_data':Sensor_data,
        'x_values': x_values.tolist(),
        'density_values': density_values.tolist(),
        'total_fish_groups': total_fish_groups,
        'total_fish_species': total_fish_species,
        'total_devices': total_devices,
        'total_categories': total_categories,
        'total_latest_fish': total_latest_fish,
        'add_num':add_num,
        'die_num':die_num,
    }
<<<<<<< Updated upstream
    return render(request, "water.html", context)
=======
    return render(request, "water.html", context)

def upload_list(request):
    # if request.method == "GET":
    return render(request,'upload1.html')



def HI_multi(request):
    "批量导入数据"
    file_object=request.FILES.get('HydrologicInfo')
    # print(type(file_object))
    if file_object:
        wb=load_workbook(file_object)
        sheet=wb.worksheets[0]
        # 循环获取数据
        for row in sheet.iter_rows(min_row=2,values_only=True):
            try:
                # 解析日期
                date_text = row[0]
                if date_text:
                    # Excel中的日期是以距离1900-01-01的天数存储的
                    date_value = datetime(1899, 12, 30) + timedelta(days=date_text)
                    date_value = date_value.date()
                else:
                    date_value = None
                expect_value=[None,'--','']
                # 读取其他列数据，示例中假设所有列都是浮点数类型
                water_quality_category = row[1] if row[1] else None
                water_temperature = row[2] if row[2]  not in expect_value else None
                pH = row[3] if row[3] not in expect_value else None
                dissolved_oxygen = row[4] if row[4] not in expect_value else None
                conductivity = row[5] if row[5] not in expect_value else None
                turbidity = row[6] if row[6] not in expect_value else None
                permanganate_index = row[7] if row[7] not in expect_value else None
                ammonia_nitrogen = row[8] if row[8] not in expect_value else None
                total_phosphorus = row[9] if row[9] not in expect_value else None
                total_nitrogen = row[10] if row[10] not in expect_value else None
                station_condition=row[11] if row[11] else None

                # 创建HydrologicInfo实例
                HydrologicInfo.objects.create(
                    monitoring_time=date_value,
                    water_quality_category=water_quality_category,
                    water_temperature=water_temperature,
                    pH=pH,
                    dissolved_oxygen=dissolved_oxygen,
                    conductivity=conductivity,
                    turbidity=turbidity,
                    permanganate_index=permanganate_index,
                    ammonia_nitrogen=ammonia_nitrogen,
                    total_phosphorus=total_phosphorus,
                    total_nitrogen=total_nitrogen,
                    station_condition=station_condition
                )
            except Exception as e:
                # 这里可以根据需要记录日志或处理异常
                print(f"Error processing row {row}: {e}")
                return errorResponse(request, '处理文件失败')

        HIsuccMsg='上传成功'
        return render(request, 'upload1.html', {
                'HIsuccMsg': HIsuccMsg
            })
    else:
        return errorResponse(request, '未传入文件')

def HI_export(request):
    # 查询数据库中所有的Fish对象
    HI_list = HydrologicInfo.objects.all()

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 写入列标题
    ws.append(['监测时间', '水质类别','水温（℃）','pH(无量纲)','溶氧量(mg/L)',	'电导率（μS/cm）','浊度（NTU）','高锰酸盐指数（mg/L）','氨氮（mg/L）','总磷（mg/L）','总氮（mg/L）','站点情况'])

    # 写入数据
    for HI in HI_list:
        ws.append([
            HI.monitoring_time,
            HI.water_quality_category,
            HI.water_temperature ,
            HI.pH ,
            HI.dissolved_oxygen,
            HI.conductivity ,
            HI.turbidity ,
            HI.permanganate_index,
            HI.ammonia_nitrogen ,
            HI.total_phosphorus ,
            HI.total_nitrogen ,
            HI.station_condition
        ])

    # 将工作簿输出到HTTP响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="HydrologicInfo_data.xlsx"'
    return response



def fish_multi(request):
    "批量导入数据"
    file_object=request.FILES.get('fish')
    # print(type(file_object))
    if file_object:
        wb=load_workbook(file_object)
        sheet=wb.worksheets[0]
        # 循环获取数据
        for row in sheet.iter_rows(min_row=2,values_only=True):
            try:
                Fish.objects.create(
                    type=row[0] if row[0] else None,
                    weight=row[1] if row[1] else 0,
                    length1=row[2] if row[2] else 0,
                    length2=row[3] if row[3] else 0,
                    length3=row[4] if row[4] else 0,
                    height=row[5] if row[5] else 0,
                    width=row[6] if row[6] else 0,
                )
            except Exception as e:
                # 这里可以根据需要记录日志或处理异常
                print(f"Error processing row {row}: {e}")
                return errorResponse(request, '处理文件失败')

        FsuccMsg='上传成功'
        return render(request, 'upload1.html', {
                'FsuccMsg': FsuccMsg
            })
    else:
        return errorResponse(request, '未传入文件')


def fish_export(request):
    # 查询数据库中所有的Fish对象
    fish_list = Fish.objects.all()

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 写入列标题
    ws.append(['Species', 'Weight(g)', 'Length1(cm)', 'Length2(cm)', 'Length3(cm)', 'Height(cm)', 'Width(cm)'])

    # 写入数据
    for fish in fish_list:
        ws.append([
            fish.type,
            fish.weight,
            fish.length1,
            fish.length2,
            fish.length3,
            fish.height,
            fish.width,
        ])

    # 将工作簿输出到HTTP响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fish_data.xlsx"'
    return response


# def staff_multi(request):
#     "批量导入数据"
#     file_object=request.FILES.get('staff')
#     # print(type(file_object))
#     if file_object:
#         wb=load_workbook(file_object)
#         sheet=wb.worksheets[0]
#         # 循环获取数据
#         for row in sheet.iter_rows(min_row=2,values_only=True):
#             try:
#                 Staff.objects.create(
#                     name=row[0] if row[0] else None,
#                     position=row[1] if row[1] else None,
#                     gender=row[2] if row[2] else None,
#                     age=row[3] if row[3] else None,
#                     phone_number=row[4] if row[4] else None,
#                     state=row[5] if row[5] else None,
#                 )
#             except Exception as e:
#                 # 这里可以根据需要记录日志或处理异常
#                 print(f"Error processing row {row}: {e}")
#                 return errorResponse(request, '处理文件失败')
#
#         SsuccMsg='上传成功'
#         return render(request, 'upload1.html', {
#                 'SsuccMsg': SsuccMsg
#             })
#     else:
#         return errorResponse(request, '未传入文件')


def device_multi(request):
    "批量导入数据"
    file_object=request.FILES.get('device')
    # print(type(file_object))
    if file_object:
        wb=load_workbook(file_object)
        sheet=wb.worksheets[0]
        # 循环获取数据
        for row in sheet.iter_rows(min_row=2,values_only=True):
            # print(row[3],' --- ',row[4],' --- ',row[5])
            try:
                # 解析日期
                # run_text = row[3]
                # if run_text:
                #     # Excel中的日期是以距离1900-01-01的天数存储的
                #     run_value = datetime(1899, 12, 30) + timedelta(days=run_text)
                #     date_value = date_value.date()
                # else:
                #     date_value = None

                Device.objects.create(
                    name=row[0] if row[0] else None,
                    category=row[6] if row[6] else None,
                    type=row[1] if row[1] else None,
                    memory=row[2] if row[2] else None,
                    run_time=row[3] if row[3] else None,
                    next_repair_time=row[4] if row[4] else None,
                    warranty_time=row[5] if row[5] else None,
                )
            except Exception as e:
                # 这里可以根据需要记录日志或处理异常
                print(f"Error processing row {row}: {e}")
                return errorResponse(request, '处理文件失败')

        DsuccMsg='上传成功'
        return render(request, 'upload1.html', {
                'DsuccMsg': DsuccMsg
            })
    else:
        return errorResponse(request, '未传入文件')

def device_export(request):
    # 查询数据库中所有的Fish对象
    device_list = Device.objects.all()

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 写入列标题
    ws.append(['名称','类型','内存大小','开始运行时间','下次维修时间',	'保修时间','种类'])

    # 写入数据
    for device in device_list:
        print(device.run_time, device.next_repair_time, device.warranty_time)
        # naive_datetime = timezone_aware_datetime.replace(tzinfo=None)
        ws.append([
            device.name ,
            device.type,
            device.memory ,
            device.run_time.replace(tzinfo=None) if device.run_time else None ,
            device.next_repair_time ,
            device.warranty_time ,
            device.category,
        ])

    # 将工作簿输出到HTTP响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="device_data.xlsx"'
    return response


def cage_multi(request):
    "批量导入数据"
    file_object=request.FILES.get('cage')
    # print(type(file_object))
    if file_object:
        wb=load_workbook(file_object)
        sheet=wb.worksheets[0]
        # 循环获取数据
        for row in sheet.iter_rows(min_row=2,values_only=True):
            # print(row[3],' --- ',row[4],' --- ',row[5])
            try:
                NetCage.objects.create(
                    width=row[0] if row[0] else None,
                    length=row[1] if row[1] else None,
                    depth=row[2] if row[2] else None,
                    longitude=row[3] if row[3] else None,
                    latitude=row[4] if row[4] else None,
                )
            except Exception as e:
                # 这里可以根据需要记录日志或处理异常
                print(f"Error processing row {row}: {e}")
                return errorResponse(request, '处理文件失败')

        CsuccMsg='上传成功'
        return render(request, 'upload1.html', {
                'CsuccMsg': CsuccMsg
            })
    else:
        return errorResponse(request, '未传入文件')

def cage_export(request):
    # 查询数据库中所有的Fish对象
    cage_list = NetCage.objects.all()

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 写入列标题
    ws.append(['宽度','长度','深度','经度','纬度'])

    # 写入数据
    for cage in cage_list:

        ws.append([
            cage.width ,
            cage.length ,
            cage.depth ,
            cage.longitude ,
            cage.latitude ,
        ])

    # 将工作簿输出到HTTP响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cage_data.xlsx"'
    return response



logger = logging.getLogger(__name__)
def get_available_dates(request):
    valid_dates = HydrologicInfo.objects.values_list('monitoring_time', flat=True).distinct()
    valid_dates = [date.strftime('%Y-%m-%d') for date in valid_dates if date is not None]
    return JsonResponse({'dates': valid_dates})


# 字段名映射
FIELD_NAME_MAPPING = {
    'ph': 'pH',
    'dissolved_oxygen': 'dissolved_oxygen',
    'turbidity': 'turbidity',
    'water_temperature': 'water_temperature',
}
logger = logging.getLogger(__name__)

def get_historical_data(request):
    date = request.GET.get('date')
    data_type = request.GET.get('dataType')

    if not date:
        logger.error('Date parameter is missing')
        return JsonResponse({'error': 'Date parameter is missing'}, status=400)

    if data_type not in FIELD_NAME_MAPPING:
        logger.error(f"Invalid data type: {data_type}")
        return JsonResponse({'error': 'Invalid data type'}, status=400)

    try:
        # 将日期格式从字符串转换为日期对象
        filtered_data = HydrologicInfo.objects.filter(monitoring_time=date)
        logger.debug(f"Filtered data for date {date}: {filtered_data}")

        if not filtered_data.exists():
            logger.error(f'No data found for the given date: {date}')
            return JsonResponse({'error': 'No data found for the given date'}, status=404)

        # 使用正确的字段名获取数据
        values = list(filtered_data.values_list(FIELD_NAME_MAPPING[data_type], flat=True))
        logger.debug(f"Values for {data_type} on {date}: {values}")

        if not values:
            logger.error(f'No values found for the given data type: {data_type} on date: {date}')
            return JsonResponse({'error': f'No values found for the given data type: {data_type}'}, status=404)

        min_value = min(values)
        max_value = max(values)
        interval_range = (max_value - min_value) / 6
        intervals = [min_value + i * interval_range for i in range(7)]

        interval_counts = [0] * 6
        for value in values:
            for i in range(6):
                if intervals[i] <= value < intervals[i + 1]:
                    interval_counts[i] += 1
                    break

        logger.debug(f"Intervals: {intervals[:-1]}, Counts: {interval_counts}")
        return JsonResponse({'intervals': intervals[:-1], 'counts': interval_counts, 'average': sum(values) / len(values)})

    except Exception as e:
        logger.error(f"Error in get_historical_data: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def get_latest_hydro_data(request):
    latest_data = HydrologicInfo.objects.latest('monitoring_time')
    data = {
        'dissolved_oxygen': latest_data.dissolved_oxygen,
        'turbidity': latest_data.turbidity,
        'ph': latest_data.pH,  
        'water_temperature': latest_data.water_temperature,
    }
    logger.debug(f"Latest hydro data: {data}")
    return JsonResponse({'data': data})

def get_latest_environment_score(request):
    latest_data = HydrologicInfo.objects.last()
    score=latest_data.calculate_environment_score()
    if latest_data:
        # weighted_score = (
        #     latest_data.water_temperature * 0.15 +
        #     latest_data.pH * 0.15 +
        #     latest_data.dissolved_oxygen * 0.15 +
        #     latest_data.conductivity * 0.1 +
        #     latest_data.turbidity * 0.1 +
        #     latest_data.permanganate_index * 0.2 +
        #     latest_data.ammonia_nitrogen * 0.15
        # )
        response_data = {
            'weighted_score': score,
            'water_temperature': latest_data.water_temperature
        }
    else:
        response_data = {
            'weighted_score': 0,
            'water_temperature': 0
        }
    return JsonResponse(response_data)

def get_latest_fish_data(request):
    latest_fish = Fish.objects.last()
    if latest_fish:
        fish_data = {
            'id': latest_fish.id,
            'type': latest_fish.type,
            'weight': latest_fish.weight,
            'length': (latest_fish.length1 + latest_fish.length2 + latest_fish.length3) / 3
        }
    else:
        fish_data = {
            'id': 'N/A',
            'type': 'N/A',
            'weight': 0,
            'length': 0
        }
    return JsonResponse(fish_data)


def get_weather_data(request):
    weather_api_url = 'https://devapi.qweather.com/v7/weather/now'
    air_quality_api_url = 'https://devapi.qweather.com/v7/air/now'
    api_key = '6213e6b3d23842548051074bbf98ee89'  # 和风天气API密钥
    location = '101310101'  # 海口市的位置ID

    params = {
        'location': location,
        'key': api_key
    }

    try:
        # 获取天气数据
        weather_response = requests.get(weather_api_url, params=params)
        weather_response.raise_for_status()  # 如果请求失败，将引发HTTPError
        weather_data = weather_response.json()

        # 获取空气质量数据
        air_quality_response = requests.get(air_quality_api_url, params=params)
        air_quality_response.raise_for_status()
        air_quality_data = air_quality_response.json()

        if 'now' in weather_data and 'now' in air_quality_data:
            result = {
                'temp': weather_data['now']['temp'],
                'windDir': weather_data['now']['windDir'],
                'windScale': weather_data['now']['windScale'],
                'humidity': weather_data['now']['humidity'],
                'aqi': air_quality_data['now']['aqi'],
                'category': air_quality_data['now']['category']
            }

            # 将AQI值转换为整数并添加AQI评估
            aqi = int(result['aqi'])
            if aqi <= 50:
                result['evaluation'] = '优'
            elif aqi <= 100:
                result['evaluation'] = '良'
            elif aqi <= 150:
                result['evaluation'] = '轻度污染'
            elif aqi <= 200:
                result['evaluation'] = '中度污染'
            elif aqi <= 300:
                result['evaluation'] = '重度污染'
            else:
                result['evaluation'] = '严重污染'

            return JsonResponse(result)
        else:
            return JsonResponse({'error': 'Unexpected response format'}, status=500)
    except requests.exceptions.RequestException as e:
        return JsonResponse({'error': str(e)}, status=500)



>>>>>>> Stashed changes
